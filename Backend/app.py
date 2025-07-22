from flask import Flask, request, jsonify, send_from_directory, send_file
from flask_cors import CORS
import mysql.connector
import jwt, datetime, secrets, smtplib
from io import BytesIO
from email.mime.text import MIMEText
import os
import traceback
from openpyxl import Workbook
from io import BytesIO
from fpdf import FPDF
from openpyxl.styles import Font, Alignment, PatternFill
from dotenv import load_dotenv
load_dotenv()
app = Flask(__name__)
CORS(app)

# === Config ===
UPLOAD_FOLDER = 'uploads'
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 100 * 1024 * 1024

app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY')
EMAIL_SENDER = os.environ.get('EMAIL_SENDER')
EMAIL_PASSWORD = os.environ.get('EMAIL_PASSWORD')

# === DB Connection ===
def get_db():
    return mysql.connector.connect(
        host=os.environ.get('DB_HOST', 'localhost'),  # Fallback to 'localhost' if not set
        user=os.environ.get('DB_USER', 'root'),
        password=os.environ.get('DB_PASSWORD', '12345'),
        database=os.environ.get('DB_NAME', 'micronetdata')
    )
def create_user_action_log():
    db = get_db()
    cur = db.cursor()
    cur.execute("""
    CREATE TABLE IF NOT EXISTS user_action_log (
      id INT AUTO_INCREMENT PRIMARY KEY,
      user_code VARCHAR(50),
      sheet_name VARCHAR(255),
      inserted INT,
      updated INT,
      timestamp DATETIME
    )""")
    db.commit()
    cur.close()
    db.close()

create_user_action_log()

# === Admin Login ===
@app.route("/admin-login", methods=["POST"])
def admin_login():
    data = request.get_json()
    username = data.get("username")
    password = data.get("password")

    db = get_db()
    cursor = db.cursor(dictionary=True)
    cursor.execute("SELECT * FROM admin WHERE username=%s AND password=%s", (username, password))
    result = cursor.fetchone()
    cursor.close()
    db.close()

    if result:
        return jsonify({"success": True, "message": "Admin login successful"})
    else:
        return jsonify({"success": False, "message": "Invalid credentials"}), 401

# === Create New User ===
@app.route("/create-user", methods=["POST"])
def create_user():
    try:
        data = request.get_json()
        if not data:
            return jsonify({"success": False, "message": "No data received"}), 400

        # Validate required fields
        required_fields = ['name', 'email', 'phone', 'password']
        for field in required_fields:
            if field not in data or not data[field]:
                return jsonify({"success": False, "message": f"Missing or empty required field: {field}"}), 400

        # Generate user code (first 3 letters of name + random 4 digits)
        user_code = f"{data['name'][:3].upper()}{secrets.randbelow(10000):04d}"

        db = get_db()
        cursor = db.cursor(dictionary=True)

        # Check if email already exists
        cursor.execute("SELECT id FROM user WHERE email = %s", (data['email'],))
        if cursor.fetchone():
            return jsonify({"success": False, "message": "Email already registered"}), 400

        # Insert new user with user_code
        cursor.execute("""
            INSERT INTO user 
            (name, email, phone, password, user_code) 
            VALUES (%s, %s, %s, %s, %s)
        """, (
            data['name'],
            data['email'],
            data['phone'],
            data['password'],
            user_code
        ))

        db.commit()
        return jsonify({
            "success": True,
            "message": "User created successfully",
            "userCode": user_code
        })

    except mysql.connector.Error as err:
        db.rollback()
        return jsonify({
            "success": False,
            "message": f"Database error: {err.msg}",
            "error_code": err.errno
        }), 500

    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"Server error: {str(e)}",
            "traceback": traceback.format_exc()
        }), 500

    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'db' in locals():
            db.close()

# === User Login with JWT ===
@app.route("/user-login", methods=["POST"])
def user_login():
    data = request.get_json()
    username, password = data.get("username"), data.get("password")

    db = get_db()
    cursor = db.cursor(dictionary=True)
    cursor.execute("SELECT * FROM user WHERE name=%s AND password=%s", (username, password))
    user = cursor.fetchone()
    cursor.close()
    db.close()

    if user:
        token = jwt.encode({
            "user_id": user["id"],
            "name": user["name"],
            "exp": datetime.datetime.utcnow() + datetime.timedelta(hours=1),
        }, app.config['SECRET_KEY'], algorithm="HS256")
        
        # Single return statement
        return jsonify({
            "success": True, 
            "token": token, 
            "name": user["name"],
            "userCode": user.get("code") or user.get("user_code") or ""  # Use whatever column name stores the code
        })
    else:
        return jsonify({"success": False, "message": "Invalid user login"}), 401  
# === Get All Users ===
@app.route('/get_users', methods=['GET'])
def get_users():
    try:
        db = get_db()
        cursor = db.cursor(dictionary=True)
        cursor.execute("SELECT id, name AS UserName, password, email, phone FROM user")
        users = cursor.fetchall()
        return jsonify(users)
    except Exception as e:
        return jsonify({"error": str(e)})
    finally:
        cursor.close()
        db.close()

# === Forgot Password ===
@app.route("/forgot-password", methods=["POST"])
def forgot_password():
    email = request.json.get("email")
    db = get_db()
    cursor = db.cursor(dictionary=True)
    cursor.execute("SELECT * FROM user WHERE email=%s", (email,))
    user = cursor.fetchone()
    if not user:
        cursor.close()
        db.close()
        return jsonify({"success": False, "message": "Email not found"}), 404

    token = secrets.token_urlsafe(32)
    cursor.execute("UPDATE user SET reset_token=%s WHERE email=%s", (token, email))
    db.commit()

    reset_link = f"http://localhost:5173/reset-password?token={token}"

    print("Password reset link:", reset_link)

    msg = MIMEText(f"Click this link to reset your password:\n\n{reset_link}")
    msg["Subject"] = "Password Reset"
    msg["From"] = EMAIL_SENDER
    msg["To"] = email

    try:
        with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
            server.login(EMAIL_SENDER, EMAIL_PASSWORD)
            server.sendmail(EMAIL_SENDER, email, msg.as_string())
        return jsonify({"success": True, "message": "Reset email sent"})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500
    finally:
        cursor.close()
        db.close()

@app.route("/reset-password", methods=["POST"])
def reset_password():
    data = request.json
    token = data.get("token")
    new_password = data.get("password")

    db = get_db()
    cursor = db.cursor(dictionary=True)
    cursor.execute("SELECT * FROM user WHERE reset_token=%s", (token,))
    user = cursor.fetchone()
    if not user:
        cursor.close()
        db.close()
        return jsonify({"success": False, "message": "Invalid or expired token"}), 400

    cursor.execute("UPDATE user SET password=%s, reset_token=NULL WHERE reset_token=%s", (new_password, token))
    db.commit()
    cursor.close()
    db.close()
    return jsonify({"success": True, "message": "Password reset successful"})

# === File Upload ===
@app.route('/upload', methods=['POST'])
def upload_file():
    if 'files' not in request.files:
        return jsonify({'success': False, 'message': 'No files part in request'}), 400

    files = request.files.getlist('files')
    saved_files = []

    db = get_db()
    cursor = db.cursor()

    for file in files:
        if file.filename == '':
            continue

        filename = file.filename
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)

        cursor.execute("INSERT INTO uploads (filename) VALUES (%s)", (filename,))
        saved_files.append(filename)

    db.commit()
    cursor.close()
    db.close()

    return jsonify({'success': True, 'uploaded': saved_files})

# === Get Uploaded Files ===
@app.route('/uploaded-files', methods=['GET'])
def get_uploaded_files():
    db = get_db()
    cursor = db.cursor(dictionary=True)
    cursor.execute("SELECT id, filename FROM uploads")
    files = cursor.fetchall()
    cursor.close()
    db.close()
    return jsonify(files)

# === Download File ===
@app.route('/download/<filename>', methods=['GET'])
def download_file(filename):
    return send_from_directory(app.config['UPLOAD_FOLDER'], filename, as_attachment=True)

# === Helper to safely read file lines ===
def safe_read_lines(filepath):
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            return f.readlines()
    except UnicodeDecodeError:
        with open(filepath, 'r', encoding='latin-1') as f:
            return f.readlines()

# === Parse Contribution Amount ===
def parse_contribution_amount(amount_str):
    return amount_str if amount_str else "0"

# === Parse ID Number ===
def parse_id_number(line):
    if len(line) >= 85:
        id_12 = line[73:85].strip()
        if id_12 and len(id_12) >= 12:
            return id_12
        return line[76:85].strip()
    return ''

# === Process Contribution File ===
def process_contribution_file(file_path):
    contributions = []
    try:
        lines = safe_read_lines(file_path)
        for line in lines:
            line = line.rstrip('\n')
            if len(line) < 23:
                continue
            try:
                record = {
                    'batch_number': line[0:7].strip(),
                    'zone_code': line[7:8].strip(),
                    'emp_number': line[8:14].strip(),
                    'cont_period': line[14:20].strip(),
                    'bank_code': line[21:22].strip() if len(line) > 21 else '',
                    'contribution': parse_contribution_amount(line[23:35].strip()),
                    'company_name': line[35:190].strip() if len(line) > 35 else ''
                }
                if record['zone_code'] and record['emp_number']:
                    contributions.append(record)
            except Exception as e:
                app.logger.warning(f"Skipping malformed contribution line: {line}\nError: {str(e)}")
    except Exception as e:
        app.logger.error(f"Error processing contribution file: {str(e)}")
        raise
    return contributions

# === Process Member File ===
def process_member_file(file_path):
    members = []
    try:
        lines = safe_read_lines(file_path)
        if not lines:
            return []
        # Skip header line (000000000000)
        lines = lines[1:] if len(lines[0].strip()) == 12 else lines

        for line in lines:
            line = line.rstrip('\n')
            if len(line) < 13:
                continue
            try:
                record = {
                    'zone_code': line[0:1].strip(),
                    'emp_number': line[1:7].strip(),
                    'member_number': line[7:13].strip(),
                    'name': line[13:72].strip(),
                    'id_number': parse_id_number(line),
                    'id_status': line[85:86].strip() if len(line) > 85 else '',
                    'mem_status': line[86:87].strip() if len(line) > 86 else ''
                }
                if record['zone_code'] and record['emp_number']:
                    members.append(record)
            except Exception as e:
                app.logger.warning(f"Skipping malformed member line: {line}\nError: {str(e)}")
    except Exception as e:
        app.logger.error(f"Error processing member file: {str(e)}")
        raise
    return members

# === Save to Dynamic Table ===
def save_to_dynamic_table(table_name, records):
    db = get_db()
    cursor = db.cursor()
    try:
        cursor.execute(f"""
            CREATE TABLE IF NOT EXISTS `{table_name}` (
                id INT AUTO_INCREMENT PRIMARY KEY,
                batch_number VARCHAR(20),
                zone_code VARCHAR(5),
                emp_number VARCHAR(10),
                cont_period VARCHAR(10),
                bank_code VARCHAR(5),
                contribution FLOAT,
                company_name TEXT,
                member_number VARCHAR(20),
                name VARCHAR(100),
                id_number VARCHAR(20),
                id_status VARCHAR(5),
                mem_status VARCHAR(5)
            )
        """)

        cursor.execute(f"DELETE FROM `{table_name}`")

        for record in records:
            cursor.execute(f"""
                INSERT INTO `{table_name}` (
                    batch_number, zone_code, emp_number, cont_period, bank_code,
                    contribution, company_name, member_number, name, id_number, id_status, mem_status
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            """, (
                record.get("batch_number", ""), record.get("zone_code", ""), record.get("emp_number", ""),
                record.get("cont_period", ""), record.get("bank_code", ""), record.get("contribution", 0.0),
                record.get("company_name", ""), record.get("member_number", ""), record.get("name", ""),
                record.get("id_number", ""), record.get("id_status", ""), record.get("mem_status", "")
            ))

        db.commit()
    except Exception as e:
        db.rollback()
        raise e
    finally:
        cursor.close()
        db.close()

# === Metadata table ===
def create_metadata_table():
    db = get_db()
    cursor = db.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS sheet_tables (
            id INT AUTO_INCREMENT PRIMARY KEY,
            sheet_name VARCHAR(255),
            zone_code VARCHAR(10),
            emp_number VARCHAR(20),
            dynamic_table_name VARCHAR(255) UNIQUE,
            batch_number VARCHAR(20),
            company_name VARCHAR(255)
        )
    """)
    db.commit()
    cursor.close()
    db.close()

create_metadata_table()

def insert_sheet_table_metadata(sheet_name, zone_code, emp_number, dynamic_table_name, batch_number, company_name):
    db = get_db()
    cursor = db.cursor()
    try:
        cursor.execute("""
            INSERT INTO sheet_tables (sheet_name, zone_code, emp_number, dynamic_table_name, batch_number, company_name)
            VALUES (%s, %s, %s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE
                batch_number = VALUES(batch_number),
                company_name = VALUES(company_name)
        """, (sheet_name, zone_code, emp_number, dynamic_table_name, batch_number, company_name))
        db.commit()
    except Exception as e:
        db.rollback()
        raise e
    finally:
        cursor.close()
        db.close()

# === Create Merged Sheet ===
@app.route('/create-merged-sheet', methods=['POST'])
def create_merged_sheet():
    data = request.get_json()
    sheet_name = data.get("sheetName")
    member_file = data.get("empMemberFile")
    contrib_file = data.get("empContributionFile")

    if not all([sheet_name, member_file, contrib_file]):
        return jsonify({"success": False, "message": "Missing fields"}), 400

    member_path = os.path.join(app.config['UPLOAD_FOLDER'], member_file)
    contrib_path = os.path.join(app.config['UPLOAD_FOLDER'], contrib_file)

    if not all(os.path.exists(p) for p in [member_path, contrib_path]):
        return jsonify({"success": False, "message": "One or more files not found"}), 400

    try:
        contributions = process_contribution_file(contrib_path)
        members = process_member_file(member_path)

        if not contributions or not members:
            return jsonify({"success": False, "message": "Empty or invalid file data"}), 400

        grouped_data = {}

        for member in members:
            key = f"{member['zone_code']}{member['emp_number']}"
            if key not in grouped_data:
                grouped_data[key] = {"members": [], "contrib": None}
            grouped_data[key]["members"].append(member)

        for contrib in contributions:
            key = f"{contrib['zone_code']}{contrib['emp_number']}"
            if key not in grouped_data:
                grouped_data[key] = {"members": [], "contrib": contrib}
            else:
                grouped_data[key]["contrib"] = contrib

        for key, group in grouped_data.items():
            contrib = group.get("contrib")
            if not contrib:
                continue

            full_records = []
            for member in group["members"]:
                merged = {**contrib, **member}
                full_records.append(merged)

            zone_code = contrib['zone_code']
            emp_number = contrib['emp_number']
            batch_number = contrib.get('batch_number', '')
            company_name = contrib.get('company_name', '')
            dynamic_table_name = f"{sheet_name}{zone_code}{emp_number}"

            save_to_dynamic_table(dynamic_table_name, full_records)
            insert_sheet_table_metadata(sheet_name, zone_code, emp_number, dynamic_table_name, batch_number, company_name)

        return jsonify({"success": True, "message": f"{len(grouped_data)} employers processed."})

    except Exception as e:
        app.logger.error(traceback.format_exc())
        return jsonify({"success": False, "message": str(e)}), 500

# === Get Employer Groups ===
@app.route('/get-employer-groups', methods=['GET'])
def get_employer_groups():
    sheet_name = request.args.get('sheetName')
    if not sheet_name:
        return jsonify({"success": False, "message": "Missing sheetName"}), 400

    db = get_db()
    cursor = db.cursor(dictionary=True)
    cursor.execute("""
        SELECT zone_code, emp_number, dynamic_table_name, batch_number, company_name
        FROM sheet_tables WHERE sheet_name=%s
    """, (sheet_name,))
    groups = cursor.fetchall()
    cursor.close()
    db.close()

    return jsonify({"success": True, "data": groups})

# === Get Members By Group ===
@app.route('/get-members-by-group', methods=['GET'])
def get_members_by_group():
    table_name = request.args.get('tableName')
    if not table_name:
        return jsonify({"success": False, "message": "Missing tableName"}), 400

    db = get_db()
    cursor = db.cursor(dictionary=True)
    try:
        cursor.execute(f"""
            SELECT member_number, name, id_number, contribution, cont_period
            FROM `{table_name}`
        """)
        members = cursor.fetchall()
    except Exception as e:
        cursor.close()
        db.close()
        return jsonify({"success": False, "message": f"Error fetching members: {str(e)}"}), 500

    cursor.close()
    db.close()
    return jsonify({"success": True, "data": members})


@app.route('/check-sheet-exists', methods=['GET'])
def check_sheet_exists():
    sheet_name = request.args.get('sheetName')
    if not sheet_name:
        return jsonify(success=False, message="Sheet name is required"), 400

    db = get_db()
    cursor = db.cursor()
    cursor.execute("SELECT COUNT(*) FROM sheet_tables WHERE sheet_name = %s", (sheet_name,))
    exists = cursor.fetchone()[0] > 0
    cursor.close()
    db.close()
    return jsonify(success=True, exists=exists)


# === Get All Sheet Names ===
@app.route('/get-sheets', methods=['GET'])
def get_sheets():
    db = get_db()
    cursor = db.cursor()
    cursor.execute("SELECT DISTINCT sheet_name FROM sheet_tables")
    sheets = [row[0] for row in cursor.fetchall()]
    cursor.close()
    db.close()
    return jsonify({"success": True, "data": sheets})


# === Get Group Labels by Sheet ===
@app.route('/get-group-labels', methods=['GET'])
def get_group_labels():
    sheet_name = request.args.get('sheetName')
    if not sheet_name:
        return jsonify({"success": False, "message": "Missing sheetName"}), 400

    db = get_db()
    cursor = db.cursor()
    cursor.execute("""
        SELECT zone_code, emp_number 
        FROM sheet_tables 
        WHERE sheet_name = %s
    """, (sheet_name,))
    results = cursor.fetchall()
    cursor.close()
    db.close()

    labels = [f"{row[0]}{row[1]}" for row in results]
    return jsonify({"success": True, "data": labels})


# === Get Members by Label (ZoneCode+EmpNumber) ===
# === Get Members by Label (ZoneCode+EmpNumber) ===
@app.route('/get-members-by-label', methods=['GET'])
@app.route('/get-members-by-label', methods=['GET'])
def get_members_by_label():
    sheet_name = request.args.get('sheetName')
    label = request.args.get('label')  # e.g. A041282

    if not sheet_name or not label or len(label) < 2:
        return jsonify({"success": False, "message": "Missing or invalid parameters"}), 400

    zone_code = label[0]
    emp_number = label[1:]

    db = get_db()
    cursor = db.cursor(dictionary=True)
    try:
        cursor.execute("""
            SELECT dynamic_table_name FROM sheet_tables 
            WHERE sheet_name = %s AND zone_code = %s AND emp_number = %s
        """, (sheet_name, zone_code, emp_number))
        row = cursor.fetchone()
        if not row:
            return jsonify({"success": False, "message": "No matching group found"}), 404

        table_name = row['dynamic_table_name']
        cursor.execute(f"""
            SELECT
                batch_number,
                zone_code,
                emp_number,
                cont_period,
                bank_code,
                contribution,
                company_name,
                member_number,
                name,
                id_number,
                id_status,
                mem_status
            FROM `{table_name}`
        """)
        members = cursor.fetchall()

    except Exception as e:
        cursor.close()
        db.close()
        return jsonify({"success": False, "message": f"DB Error: {str(e)}"}), 500

    cursor.close()
    db.close()
    return jsonify({"success": True, "data": members})

def create_mismatch_table():
    db = get_db()
    cursor = db.cursor()
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS batchnumber_mismatch (
            id INT AUTO_INCREMENT PRIMARY KEY,
            sheet_name VARCHAR(255),
            batch_number VARCHAR(20),
            zone_code VARCHAR(5),
            emp_number VARCHAR(10),
            formatone_total DECIMAL(15,2),
            merged_sheet_total DECIMAL(15,2),
            user_code VARCHAR(50),
            mismatch_date DATETIME DEFAULT CURRENT_TIMESTAMP,
            resolved BOOLEAN DEFAULT FALSE
        )
    """)
    db.commit()
    cursor.close()
    db.close()

create_mismatch_table()

@app.route('/get-mismatches', methods=['GET'])
def get_mismatches():
    db = get_db()
    cursor = db.cursor(dictionary=True)
    
    try:
        cursor.execute("""
            SELECT * FROM batchnumber_mismatch 
            WHERE resolved = FALSE
            ORDER BY mismatch_date DESC
        """)
        mismatches = cursor.fetchall()
        return jsonify({"success": True, "data": mismatches})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500
    finally:
        cursor.close()
        db.close()
        
# === Delete Sheet ===
@app.route('/delete-sheet', methods=['DELETE'])
def delete_sheet():
    sheet_name = request.args.get('sheetName')
    if not sheet_name:
        return jsonify({"success": False, "message": "Sheet name is required"}), 400

    db = get_db()
    cursor = db.cursor()
    try:
        # Get all dynamic tables associated with this sheet
        cursor.execute("""
            SELECT dynamic_table_name FROM sheet_tables 
            WHERE sheet_name = %s
        """, (sheet_name,))
        tables = cursor.fetchall()

        # Drop each dynamic table
        for table in tables:
            cursor.execute(f"DROP TABLE IF EXISTS `{table[0]}`")

        # Delete metadata entries
        cursor.execute("""
            DELETE FROM sheet_tables 
            WHERE sheet_name = %s
        """, (sheet_name,))

        db.commit()
        return jsonify({"success": True, "message": f"Sheet '{sheet_name}' deleted successfully."})

    except Exception as e:
        db.rollback()
        return jsonify({"success": False, "message": str(e)}), 500
    finally:
        cursor.close()
        db.close()


@app.route('/save-format-one', methods=['POST'])
def save_format_one():
    db = None
    cur = None
    try:
        data = request.get_json()
        if not data:
            return jsonify(success=False, message="No data received"), 400

        required_fields = ['sheetName', 'userCode', 'rows']
        for field in required_fields:
            if field not in data:
                return jsonify(success=False, message=f"Missing required field: {field}"), 400

        sheet_name = data['sheetName']
        rows = data['rows']
        user_code = data['userCode']

        if not isinstance(rows, list):
            return jsonify(success=False, message="Rows should be an array"), 400

        # Validate each row has required fields
        required_row_fields = ['batchNumber', 'zoneCode', 'empNumber', 'contributed', 'membersNumber', 'contribution']
        for i, row in enumerate(rows):
            for field in required_row_fields:
                if field not in row:
                    return jsonify(
                        success=False,
                        message=f"Row {i+1} missing required field: {field}"
                    ), 400

        # Calculate frontend total
        frontend_total = sum(float(row.get('contribution', 0)) for row in rows)
        
        # Get the first row's batch number, zone code and emp number for comparison
        first_row = rows[0]
        batch_number = first_row['batchNumber']
        zone_code = first_row['zoneCode']
        emp_number = first_row['empNumber']
        
        # Get the dynamic table name for this group
        dynamic_table_name = f"{sheet_name}{zone_code}{emp_number}"
        
        # Calculate backend total from the dynamic table (single row's contribution)
        backend_total = 0
        try:
            db = get_db()
            cur = db.cursor(dictionary=True)
            
            # Check if dynamic table exists
            cur.execute("SHOW TABLES LIKE %s", (dynamic_table_name,))
            if cur.fetchone():
                # Get the specific row's contribution from dynamic table
                cur.execute(f"""
                    SELECT contribution 
                    FROM `{dynamic_table_name}`
                    WHERE batch_number = %s 
                    AND zone_code = %s 
                    AND emp_number = %s
                    LIMIT 1
                """, (batch_number, zone_code, emp_number))
                result = cur.fetchone()
                if result:
                    backend_total = float(result['contribution']) if result['contribution'] else 0
        except Exception as e:
            app.logger.error(f"Error calculating backend total: {str(e)}")
        
        # Compare totals (ignoring decimals)
        mismatch_detected = int(frontend_total) != int(backend_total)
        
        table_name = f"{sheet_name}Cfile"
        
        # Create table if not exists with proper schema
        cur.execute(f"""
        CREATE TABLE IF NOT EXISTS `{table_name}` (
          id INT AUTO_INCREMENT PRIMARY KEY,
          batch_number VARCHAR(7) NOT NULL,
          zone_code VARCHAR(1) NOT NULL,
          emp_number VARCHAR(6) NOT NULL,
          cont_period VARCHAR(6) NOT NULL,
          member_number VARCHAR(6) NOT NULL,
          record_id VARCHAR(2) DEFAULT '5',
          page_no VARCHAR(4),
          contribution CHAR(11) NOT NULL,
          modified_by VARCHAR(50),
          modified_at DATETIME DEFAULT CURRENT_TIMESTAMP,
          UNIQUE KEY unique_member (batch_number, zone_code, emp_number, cont_period, member_number)
        )""")
        db.commit()

        inserted = updated = 0
        now = datetime.datetime.now()
        mismatches = 0

        for row in rows:
            # Format data according to requirements
            batch_number = str(row.get('batchNumber', '')).strip().zfill(7)
            zone_code = str(row.get('zoneCode', '')).strip().upper()
            emp_number = str(row.get('empNumber', '')).strip().zfill(6)
            cont_period = str(row.get('contributed', '')).strip().zfill(6)
            member_number = str(row.get('membersNumber', '')).strip().zfill(6)
            page_no = str(row.get('pageNo', '0')).strip().zfill(4)
            contribution = str(row.get('contribution', '0')).strip().zfill(11)
            
            # Check if record exists with the composite key
            cur.execute(f"""
            SELECT 
                page_no as existing_page_no,
                contribution as existing_contribution
            FROM `{table_name}` 
            WHERE 
                batch_number=%s AND 
                zone_code=%s AND 
                emp_number=%s AND 
                cont_period=%s AND 
                member_number=%s
            """, (batch_number, zone_code, emp_number, cont_period, member_number))
            
            existing_record = cur.fetchone()
            
            if existing_record:
                # Check if any data needs updating
                needs_update = (
                    existing_record['existing_page_no'] != page_no or
                    existing_record['existing_contribution'] != contribution
                )
                
                if needs_update:
                    # Update existing record
                    cur.execute(f"""
                    UPDATE `{table_name}` SET 
                        page_no=%s,
                        contribution=%s,
                        modified_by=%s,
                        modified_at=%s
                    WHERE 
                        batch_number=%s AND 
                        zone_code=%s AND 
                        emp_number=%s AND 
                        cont_period=%s AND 
                        member_number=%s
                    """, (
                        page_no, contribution, user_code, now,
                        batch_number, zone_code, emp_number, cont_period, member_number
                    ))
                    updated += 1
            else:
                # Insert new record
                cur.execute(f"""
                INSERT INTO `{table_name}` (
                    batch_number, zone_code, emp_number, cont_period,
                    member_number, record_id, page_no, contribution,
                    modified_by, modified_at
                ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (
                    batch_number, zone_code, emp_number, cont_period,
                    member_number, '5', page_no, contribution,
                    user_code, now
                ))
                inserted += 1

        # Create mismatch table if totals don't match
        if mismatch_detected:
            mismatch_table_name = f"{batch_number}_mismatches"
            
            try:
                # Create mismatch table if not exists
                cur.execute(f"""
                CREATE TABLE IF NOT EXISTS `{mismatch_table_name}` (
                    id INT AUTO_INCREMENT PRIMARY KEY,
                    batch_number VARCHAR(7) NOT NULL,
                    zone_code VARCHAR(1) NOT NULL,
                    emp_number VARCHAR(6) NOT NULL,
                    frontend_total DECIMAL(15,2) NOT NULL,
                    backend_total DECIMAL(15,2) NOT NULL,
                    difference DECIMAL(15,2) NOT NULL,
                    user_code VARCHAR(50) NOT NULL,
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    UNIQUE KEY unique_mismatch (batch_number, zone_code, emp_number)
                )
                """)
                
                # Insert mismatch record
                difference = frontend_total - backend_total
                cur.execute(f"""
                INSERT INTO `{mismatch_table_name}` (
                    batch_number, zone_code, emp_number,
                    frontend_total, backend_total, difference, user_code
                ) VALUES (%s, %s, %s, %s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE
                    frontend_total = VALUES(frontend_total),
                    backend_total = VALUES(backend_total),
                    difference = VALUES(difference),
                    user_code = VALUES(user_code),
                    created_at = CURRENT_TIMESTAMP
                """, (
                    batch_number, zone_code, emp_number,
                    round(frontend_total, 2), round(backend_total, 2), 
                    round(difference, 2), user_code
                ))
                
                mismatches = 1
                db.commit()
            except Exception as e:
                app.logger.error(f"Error creating mismatch table: {str(e)}")
                db.rollback()

        # Log the user action
        cur.execute("""
        INSERT INTO user_action_log 
        (user_code, sheet_name, inserted, updated, timestamp)
        VALUES (%s, %s, %s, %s, %s)
        """, (user_code, sheet_name, inserted, updated, now))
        
        db.commit()
        
        return jsonify(
            success=True, 
            message="Data processed successfully",
            inserted=inserted, 
            updated=updated,
            mismatches=mismatches,
            frontend_total=round(frontend_total, 2),
            backend_total=round(backend_total, 2)
        )

    except mysql.connector.Error as db_err:
        if db:
            db.rollback()
        return jsonify(
            success=False,
            message=f"Database error: {str(db_err)}",
            error_code=db_err.errno
        ), 500
        
    except Exception as e:
        if db:
            db.rollback()
        return jsonify(
            success=False,
            message=f"Server error: {str(e)}"
        ), 500
        
    finally:
        if cur:
            cur.close()
        if db:
            db.close()        

@app.route('/download-cfile', methods=['GET'])
def download_cfile():
    sheet_name = request.args.get('sheetName')
    if not sheet_name:
        return jsonify(success=False, message="Sheet name is required"), 400

    db = None
    cur = None
    try:
        db = get_db()
        cur = db.cursor(dictionary=True)
        
        table_name = f"{sheet_name}Cfile"
        
        # Check if table exists
        cur.execute("SHOW TABLES LIKE %s", (table_name,))
        if not cur.fetchone():
            return jsonify(success=False, message=f"Table {table_name} does not exist"), 404
        
        # Get all data from the table
        cur.execute(f"""
        SELECT 
            batch_number,
            zone_code,
            emp_number,
            cont_period,
            member_number,
            record_id,
            page_no,
            contribution
        FROM `{table_name}`
        ORDER BY batch_number, zone_code, emp_number, member_number
        """)
        
        file_data = cur.fetchall()
        
        return jsonify(
            success=True,
            message="Data retrieved successfully",
            fileData=file_data
        )
        
    except Exception as e:
        return jsonify(
            success=False,
            message=f"Server error: {str(e)}"
        ), 500
    finally:
        if cur:
            cur.close()
        if db:
            db.close()

@app.route('/get-cfile-row-count', methods=['GET'])
def get_cfile_row_count():
    sheet_name = request.args.get('sheetName')
    if not sheet_name:
        return jsonify(success=False, message="Sheet name is required"), 400

    db = None
    cur = None
    try:
        db = get_db()
        cur = db.cursor(dictionary=True)
        
        table_name = f"{sheet_name}Cfile"
        
        # Check if table exists
        cur.execute("SHOW TABLES LIKE %s", (table_name,))
        if not cur.fetchone():
            return jsonify(success=False, message=f"Table {table_name} does not exist"), 404
        
        # Get row count
        cur.execute(f"SELECT COUNT(*) as row_count FROM `{table_name}`")
        result = cur.fetchone()
        
        return jsonify(
            success=True,
            rowCount=result['row_count']
        )
        
    except Exception as e:
        return jsonify(
            success=False,
            message=f"Error getting row count: {str(e)}"
        ), 500
    finally:
        if cur:
            cur.close()
        if db:
            db.close()
@app.route('/get-mismatch-batches', methods=['GET'])
def get_mismatch_batches():
    db = None
    cur = None
    try:
        db = get_db()
        cur = db.cursor(dictionary=True)
        
        # Get all tables with _mismatches suffix and their record counts
        cur.execute("""
            SELECT 
                REPLACE(table_name, '_mismatches', '') AS batch_number,
                (SELECT COUNT(*) FROM `{table_name}`) AS mismatch_count
            FROM information_schema.tables
            WHERE table_schema = DATABASE() 
            AND table_name LIKE '%\\_mismatches' ESCAPE '\\'
        """)
        
        batches = cur.fetchall()
        return jsonify(success=True, data=batches)
        
    except Exception as e:
        app.logger.error(f"Error fetching mismatch batches: {str(e)}")
        return jsonify(success=False, message=str(e)), 500
    finally:
        if cur: cur.close()
        if db: db.close()

@app.route('/get-mismatch-details', methods=['GET'])
def get_mismatch_details():
    batch_number = request.args.get('batchNumber')
    if not batch_number:
        return jsonify(success=False, message="Batch number required"), 400
    
    # Remove _mismatches suffix if present
    if batch_number.endswith('_mismatches'):
        batch_number = batch_number.replace('_mismatches', '')
    
    db = None
    cur = None
    try:
        db = get_db()
        cur = db.cursor(dictionary=True)
        
        table_name = f"{batch_number}_mismatches"
        
        # Check if table exists
        cur.execute("SHOW TABLES LIKE %s", (table_name,))
        if not cur.fetchone():
            return jsonify(success=False, message="Mismatch table not found"), 404
        
        # Get mismatch details
        cur.execute(f"""
            SELECT 
                batch_number,
                zone_code,
                emp_number,
                frontend_total,
                backend_total,
                difference,
                user_code,
                created_at
            FROM `{table_name}`
            LIMIT 1
        """)
        
        details = cur.fetchone()
        if not details:
            return jsonify(success=False, message="No mismatch records found"), 404
            
        return jsonify(success=True, data=details)
        
    except Exception as e:
        app.logger.error(f"Error fetching mismatch details: {str(e)}")
        return jsonify(success=False, message=str(e)), 500
    finally:
        if cur: cur.close()
        if db: db.close()

@app.route('/download-mismatch-pdf', methods=['GET'])
def download_mismatch_pdf():
    batch_number = request.args.get('batchNumber')
    if not batch_number:
        return jsonify({"success": False, "message": "Batch number is required"}), 400

    # Remove _mismatches suffix if present
    if batch_number.endswith('_mismatches'):
        batch_number = batch_number.replace('_mismatches', '')

    try:
        db = get_db()
        cursor = db.cursor(dictionary=True)

        table_name = f"{batch_number}_mismatches"
        
        # Check if table exists
        cursor.execute("""
            SELECT COUNT(*) as table_exists 
            FROM information_schema.tables 
            WHERE table_schema = DATABASE() AND table_name = %s
        """, (table_name,))
        
        if not cursor.fetchone()['table_exists']:
            return jsonify({"success": False, "message": f"No mismatch data found for batch {batch_number}"}), 404

        # Get all rows
        cursor.execute(f"""
            SELECT batch_number, zone_code, emp_number, 
                   frontend_total, backend_total, difference,
                   user_code
            FROM `{table_name}`
        """)
        rows = cursor.fetchall()

        if not rows:
            return jsonify({"success": False, "message": "No data found in mismatch report"}), 404

        # Create PDF
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)

        # Title
        pdf.cell(200, 10, txt=f"Mismatch Report - Batch: {batch_number}", ln=1, align='C')
        pdf.ln(5)
        pdf.cell(200, 10, txt=f"Generated on: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", ln=1)
        pdf.ln(10)

        # Table headers (removed Date column)
        headers = ["Batch", "Zone", "Emp No", "Frontend Total", "Backend Total", "Difference", "Reported By"]
        col_widths = [20, 15, 20, 30, 30, 25, 25]

        # Header style
        pdf.set_fill_color(200, 220, 255)
        for i, header in enumerate(headers):
            pdf.cell(col_widths[i], 10, header, 1, 0, 'C', 1)
        pdf.ln()

        # Table rows
        pdf.set_fill_color(255, 255, 255)
        for row in rows:
            # Highlight rows with difference
            if row['difference'] != 0:
                pdf.set_fill_color(255, 200, 200)  # Light red for mismatches
            else:
                pdf.set_fill_color(200, 255, 200)  # Light green for matches
            
            pdf.cell(col_widths[0], 10, row['batch_number'], 1, 0, 'C', 1)
            pdf.cell(col_widths[1], 10, row['zone_code'], 1, 0, 'C', 1)
            pdf.cell(col_widths[2], 10, row['emp_number'], 1, 0, 'C', 1)
            pdf.cell(col_widths[3], 10, f"{row['frontend_total']:.2f}", 1, 0, 'R', 1)
            pdf.cell(col_widths[4], 10, f"{row['backend_total']:.2f}", 1, 0, 'R', 1)
            pdf.cell(col_widths[5], 10, f"{row['difference']:.2f}", 1, 0, 'R', 1)
            pdf.cell(col_widths[6], 10, row['user_code'], 1, 0, 'C', 1)
            pdf.ln()

        # Create PDF in memory
        pdf_bytes = pdf.output(dest='S').encode('latin-1')
        pdf_io = BytesIO(pdf_bytes)

        return send_file(
            pdf_io,
            as_attachment=True,
            download_name=f"Mismatch_Report_{batch_number}.pdf",
            mimetype="application/pdf"
        )

    except Exception as e:
        app.logger.error(f"PDF report error: {str(e)}\n{traceback.format_exc()}")
        return jsonify({
            "success": False,
            "message": "Failed to generate PDF report",
            "error": str(e)
        }), 500
    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'db' in locals():
            db.close()

@app.route('/download-mismatch-excel', methods=['GET'])
def download_mismatch_excel():
    batch_number = request.args.get('batchNumber')
    if not batch_number:
        return jsonify({"success": False, "message": "Batch number is required"}), 400

    # Remove _mismatches suffix if present
    if batch_number.endswith('_mismatches'):
        batch_number = batch_number.replace('_mismatches', '')

    try:
        db = get_db()
        cursor = db.cursor(dictionary=True)

        table_name = f"{batch_number}_mismatches"
        
        # Check if table exists
        cursor.execute("""
            SELECT COUNT(*) as table_exists 
            FROM information_schema.tables 
            WHERE table_schema = DATABASE() AND table_name = %s
        """, (table_name,))
        
        if not cursor.fetchone()['table_exists']:
            return jsonify({"success": False, "message": f"No mismatch data found for batch {batch_number}"}), 404

        # Get all rows (removed created_at from query)
        cursor.execute(f"""
            SELECT batch_number, zone_code, emp_number, 
                   frontend_total, backend_total, difference,
                   user_code
            FROM `{table_name}`
        """)
        rows = cursor.fetchall()

        if not rows:
            return jsonify({"success": False, "message": "No data found in mismatch report"}), 404

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Mismatch Report"

        # Headers (removed Date column)
        headers = ["Batch", "Zone", "Emp No", "Frontend Total", "Backend Total", "Difference", "Reported By"]
        ws.append(headers)

        # Style headers
        header_fill = PatternFill(start_color="C9D9F2", end_color="C9D9F2", fill_type="solid")
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
            cell.fill = header_fill

        # Define fill patterns
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

        # Add data rows
        for row in rows:
            ws.append([
                row['batch_number'],
                row['zone_code'],
                row['emp_number'],
                float(row['frontend_total']),
                float(row['backend_total']),
                float(row['difference']),
                row['user_code']
            ])

        # Apply conditional formatting
        for row in ws.iter_rows(min_row=2, max_row=len(rows)+1, min_col=6, max_col=6):
            for cell in row:
                if cell.value < 0:
                    cell.fill = red_fill
                elif cell.value > 0:
                    cell.fill = green_fill
                else:
                    cell.fill = white_fill
                cell.number_format = '0.00'

        # Format number columns
        for col in [4, 5, 6]:  # Frontend, Backend, Difference columns
            for row in ws.iter_rows(min_row=2, max_row=len(rows)+1, min_col=col, max_col=col):
                for cell in row:
                    cell.number_format = '0.00'

        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width

        # Send file
        excel_io = BytesIO()
        wb.save(excel_io)
        excel_io.seek(0)

        return send_file(
            excel_io,
            as_attachment=True,
            download_name=f"Mismatch_Report_{batch_number}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        app.logger.error(f"Excel report error: {str(e)}\n{traceback.format_exc()}")
        return jsonify({
            "success": False,
            "message": "Failed to generate Excel report",
            "error": str(e)
        }), 500
    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'db' in locals():
            db.close()
    # === Save Format Two Data ===
@app.route("/save-format-two", methods=["POST"])
def save_format_two():
    db = None
    cur = None
    try:
        data = request.get_json()
        if not data:
            return jsonify(success=False, message="No data received"), 400

        required_fields = ['sheetName', 'userCode', 'rows']
        for field in required_fields:
            if field not in data:
                return jsonify(success=False, message=f"Missing required field: {field}"), 400

        sheet_name = data['sheetName']
        rows = data['rows']
        user_code = data['userCode']
        table_name = f"{sheet_name}Mfile"

        db = get_db()
        cur = db.cursor(dictionary=True)

        # Create table if not exists
        cur.execute(f"""
        CREATE TABLE IF NOT EXISTS `{table_name}` (
          id INT AUTO_INCREMENT PRIMARY KEY,
          batch_number VARCHAR(8) NOT NULL,
          zone_code VARCHAR(1) NOT NULL,
          emp_number VARCHAR(6) NOT NULL,
          member_number VARCHAR(6) NOT NULL,
          last_name VARCHAR(60),
          initials VARCHAR(10),
          id_number VARCHAR(12),
          id_status VARCHAR(1),
          mem_status VARCHAR(1),
          operation_code VARCHAR(1),
          full_name VARCHAR(100),
          modified_by VARCHAR(50),
          modified_at DATETIME DEFAULT CURRENT_TIMESTAMP,
          UNIQUE KEY unique_member (batch_number, zone_code, emp_number, member_number)
        )
        """)
        db.commit()

        inserted = updated = 0
        now = datetime.datetime.now()

        for row in rows:
            # Extract and format data
            batch_number = str(row.get('batchNumber', '')).strip().ljust(8)[:8]
            zone_code = str(row.get('zoneCode', '')).strip().ljust(1)[:1]
            emp_number = str(row.get('empNumber', '')).strip().ljust(6)[:6]
            member_number = str(row.get('memNumber', '')).strip().ljust(6)[:6]
            last_name = str(row.get('lastName', '')).strip().ljust(60)[:60]
            initials = str(row.get('initials', '')).strip().ljust(10)[:10]  # NEW FIELD
            id_number = str(row.get('idNumber', '')).strip()
            if not id_number:
                id_number = '000000000000'
            elif len(id_number) == 9:
                id_number = '000' + id_number
                id_number = id_number.ljust(12)[:12]

            id_status = str(row.get('idStatus', '')).strip().ljust(1)[:1]
            mem_status = str(row.get('memStatus', '')).strip().ljust(1)[:1]
            operation_code = str(row.get('operationCode', '')).strip().ljust(1)[:1]
            full_name = str(row.get('fullName', '')).strip().ljust(100)[:100]

            # Check if record exists
            cur.execute(f"""
            SELECT id FROM `{table_name}` 
            WHERE 
                batch_number=%s AND 
                zone_code=%s AND 
                emp_number=%s AND 
                member_number=%s
            """, (batch_number, zone_code, emp_number, member_number))

            exists = cur.fetchone()

            if exists:
                # Update existing record
                cur.execute(f"""
                UPDATE `{table_name}` SET 
                    last_name=%s,
                    initials=%s,
                    id_number=%s,
                    id_status=%s,
                    mem_status=%s,
                    operation_code=%s,
                    full_name=%s,
                    modified_by=%s,
                    modified_at=%s
                WHERE id=%s
                """, (
                    last_name, initials, id_number, id_status, mem_status,
                    operation_code, full_name, user_code, now,
                    exists['id']
                ))
                updated += 1
            else:
                # Insert new record
                cur.execute(f"""
                INSERT INTO `{table_name}` (
                    batch_number, zone_code, emp_number, member_number,
                    last_name, initials, id_number, id_status, mem_status,
                    operation_code, full_name, modified_by
                ) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                """, (
                    batch_number, zone_code, emp_number, member_number,
                    last_name, initials, id_number, id_status, mem_status,
                    operation_code, full_name, user_code
                ))
                inserted += 1

        # Log the user action
        cur.execute("""
        INSERT INTO user_action_log 
        (user_code, sheet_name, inserted, updated, timestamp)
        VALUES (%s, %s, %s, %s, %s)
        """, (user_code, sheet_name, inserted, updated, now))
        db.commit()

        return jsonify(
            success=True,
            message="Data saved successfully",
            inserted=inserted,
            updated=updated
        )

    except Exception as e:
        if db:
            db.rollback()
        return jsonify(success=False, message=str(e)), 500

    finally:
        if cur:
            cur.close()
        if db:
            db.close()



@app.route('/get-mfile-row-count', methods=['GET'])
def get_mfile_row_count():
    sheet_name = request.args.get('sheetName')
    if not sheet_name:
        return jsonify(success=False, message="Sheet name is required"), 400

    db = None
    cur = None
    try:
        db = get_db()
        cur = db.cursor(dictionary=True)
        
        table_name = f"{sheet_name}Mfile"
        
        # Check if table exists
        cur.execute("SHOW TABLES LIKE %s", (table_name,))
        if not cur.fetchone():
            return jsonify(success=False, message=f"Table {table_name} does not exist"), 404
        
        # Get row count
        cur.execute(f"SELECT COUNT(*) as row_count FROM `{table_name}`")
        result = cur.fetchone()
        
        return jsonify(
            success=True,
            rowCount=result['row_count']
        )
        
    except Exception as e:
        return jsonify(
            success=False,
            message=f"Error getting row count: {str(e)}"
        ), 500
    finally:
        if cur:
            cur.close()
        if db:
            db.close()


# === Download Mfile ===
@app.route('/download-mfile', methods=['GET'])
def download_mfile():
    sheet_name = request.args.get('sheetName')
    if not sheet_name:
        return jsonify(success=False, message="Sheet name is required"), 400

    db = None
    cur = None
    try:
        db = get_db()
        cur = db.cursor(dictionary=True)

        table_name = f"{sheet_name}Mfile"

        # Check if table exists
        cur.execute("SHOW TABLES LIKE %s", (table_name,))
        if not cur.fetchone():
            return jsonify(success=False, message=f"Table {table_name} not found"), 404

        # Get all data
        cur.execute(f"""
        SELECT 
            batch_number,
            zone_code,
            emp_number,
            member_number,
            last_name,
            initials,
            id_number,
            id_status,
            mem_status,
            operation_code,
            full_name
        FROM `{table_name}`
        ORDER BY batch_number, zone_code, emp_number, member_number
        """)

        records = cur.fetchall()

        # Get row count
        cur.execute(f"SELECT COUNT(*) as row_count FROM `{table_name}`")
        row_count = cur.fetchone()['row_count']

        # Format each line with exact field lengths and positions
        file_lines = []
        for record in records:
            line = (
                str(record['batch_number']).strip().ljust(7)[:7] +     # 1-7
                str(record['zone_code']).strip().ljust(1)[:1] +        # 8
                str(record['emp_number']).strip().ljust(6)[:6] +       # 9-14
                str(record['member_number']).strip().ljust(6)[:6] +    # 15-20
                str(record['last_name']).strip().ljust(60)[:60] +      # 21-80
                str(record['initials']).strip().ljust(20)[:20] +       # 81-100
                str(record['id_number']).strip().ljust(12)[:12] +      # 101-112
                str(record['id_status']).strip().ljust(1)[:1] +        # 113
                str(record['mem_status']).strip().ljust(1)[:1] +       # 114
                str(record['operation_code']).strip().ljust(1)[:1] +   # 115
                str(record['full_name']).strip().ljust(100)[:100]      # 116-215
            )
            file_lines.append(line)

        return jsonify(
            success=True,
            fileData=file_lines,
            fileName=f"{sheet_name}Mfile.txt",
            rowCount=row_count
        )

    except Exception as e:
        return jsonify(success=False, message=str(e)), 500
    finally:
        if cur: cur.close()
        if db: db.close()
        
@app.route('/get-company-address', methods=['GET'])
def get_company_address():
    batch_number = request.args.get('batchNumber')
    zone_code = request.args.get('zoneCode')
    
    if not batch_number or not zone_code:
        return jsonify({"success": False, "message": "Missing parameters"}), 400

    db = get_db()
    cursor = db.cursor(dictionary=True)
    
    try:
        cursor.execute("""
            SELECT company_name 
            FROM sheet_tables 
            WHERE batch_number LIKE %s AND zone_code = %s
            LIMIT 1
        """, (f"%{batch_number}%", zone_code))
        
        result = cursor.fetchone()
        
        if result and result.get('company_name'):
            return jsonify({
                "success": True,
                "address": result['company_name']
            })
            
        return jsonify({
            "success": True,
            "address": "Address not available"
        })
        
    except Exception as e:
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500
    finally:
        cursor.close()
        db.close()

# Add this new endpoint to your Flask app
@app.route('/save-volume-counts', methods=['POST'])
def save_volume_counts():
    data = request.get_json()
    sheet_name = data.get('sheetName')
    user_code = data.get('userCode')
    counts = data.get('counts')  # {'NN': count, 'OI': count, 'MS': count}

    if not all([sheet_name, user_code, counts]):
        return jsonify({"success": False, "message": "Missing required fields"}), 400

    db = get_db()
    cursor = db.cursor()
    
    try:
        # Create dedicated volume counts table
        cursor.execute(f"""
        CREATE TABLE IF NOT EXISTS `{sheet_name}_volume_report` (
          id INT AUTO_INCREMENT PRIMARY KEY,
          user_code VARCHAR(50) NOT NULL,
          NN_count INT DEFAULT 0,
          OI_count INT DEFAULT 0,
          MS_count INT DEFAULT 0,
          timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
          UNIQUE KEY unique_user (user_code)
        )
        """)

        # Insert or update counts
        cursor.execute(f"""
        INSERT INTO `{sheet_name}_volume_report` 
        (user_code, NN_count, OI_count, MS_count)
        VALUES (%s, %s, %s, %s)
        ON DUPLICATE KEY UPDATE
          NN_count = NN_count + VALUES(NN_count),
          OI_count = OI_count + VALUES(OI_count),
          MS_count = MS_count + VALUES(MS_count)
        """, (
            user_code,
            counts.get('NN', 0),
            counts.get('OI', 0),
            counts.get('MS', 0)
        ))

        db.commit()
        return jsonify({"success": True, "message": "Volume counts updated"})
        
    except Exception as e:
        db.rollback()
        return jsonify({"success": False, "message": str(e)}), 500
    finally:
        cursor.close()
        db.close()
@app.route('/get-batches', methods=['GET'])
def get_batches():
    try:
        db = get_db()
        cursor = db.cursor(dictionary=True)
        
        cursor.execute("""
            SELECT DISTINCT batch_number 
            FROM sheet_tables 
            WHERE batch_number IS NOT NULL AND batch_number != ''
            ORDER BY batch_number
        """)
        
        batches = [row['batch_number'] for row in cursor.fetchall()]
        
        return jsonify({
            "success": True,
            "data": batches
        })
        
    except Exception as e:
        return jsonify({
            "success": False,
            "message": str(e)
        }), 500
    finally:
        cursor.close()
        db.close()

@app.route('/download-volume-pdf', methods=['GET'])
def download_volume_pdf():
    sheet_name = request.args.get('sheetName')
    if not sheet_name:
        return jsonify({"success": False, "message": "Sheet name is required"}), 400

    try:
        db = get_db()
        cursor = db.cursor(dictionary=True)

        # Check if table exists
        table_name = f"{sheet_name}_volume_report"
        cursor.execute("""
            SELECT COUNT(*) as table_exists 
            FROM information_schema.tables 
            WHERE table_schema = DATABASE() AND table_name = %s
        """, (table_name,))
        table_exists = cursor.fetchone()['table_exists']
        if not table_exists:
            return jsonify({"success": False, "message": "No volume report table found for this sheet"}), 404

        # Get all rows
        cursor.execute(f"""
            SELECT user_code, NN_count, OI_count, MS_count 
            FROM `{table_name}`
        """)
        rows = cursor.fetchall()

        if not rows:
            return jsonify({"success": False, "message": "No data found in volume report"}), 404

        # Totals
        total_nn = sum(row['NN_count'] or 0 for row in rows)
        total_oi = sum(row['OI_count'] or 0 for row in rows)
        total_ms = sum(row['MS_count'] or 0 for row in rows)

        # Create PDF
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)

        pdf.cell(200, 10, txt=f"Volume Report - Sheet: {sheet_name}", ln=1, align='C')
        pdf.ln(5)
        pdf.cell(200, 10, txt=f"Report Date: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}", ln=1)
        pdf.ln(10)

        # Table headers
        headers = [ "User Code","Input of Name & ID Number (NN)","Input ID Number (OI)","Change of Member Status (MS)"]

        col_widths = [25, 65, 40, 60]

        pdf.set_fill_color(200, 220, 255)
        for i, header in enumerate(headers):
            pdf.cell(col_widths[i], 10, header, 1, 0, 'C', 1)
        pdf.ln()

        # Table rows
        for row in rows:
            pdf.cell(col_widths[0], 10, row['user_code'], 1)
            pdf.cell(col_widths[1], 10, str(row['NN_count']), 1, 0, 'C')
            pdf.cell(col_widths[2], 10, str(row['OI_count']), 1, 0, 'C')
            pdf.cell(col_widths[3], 10, str(row['MS_count']), 1, 0, 'C')
            pdf.ln()

        # Totals
        pdf.set_font("Arial", 'B', 12)
        pdf.cell(col_widths[0], 10, "TOTAL", 1)
        pdf.cell(col_widths[1], 10, str(total_nn), 1, 0, 'C')
        pdf.cell(col_widths[2], 10, str(total_oi), 1, 0, 'C')
        pdf.cell(col_widths[3], 10, str(total_ms), 1, 0, 'C')
        pdf.ln()

        # Create PDF in memory
        pdf_bytes = pdf.output(dest='S').encode('latin-1')
        pdf_io = BytesIO(pdf_bytes)

        return send_file(
            pdf_io,
            as_attachment=True,
            download_name=f"Volume_Report_{sheet_name}.pdf",
            mimetype="application/pdf"
        )

    except Exception as e:
        app.logger.error(f"PDF report error: {str(e)}\n{traceback.format_exc()}")
        return jsonify({
            "success": False,
            "message": "Failed to generate PDF report",
            "error": str(e)
        }), 500
    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'db' in locals():
            db.close()
@app.route('/download-volume-excel', methods=['GET'])
def download_volume_excel():
    sheet_name = request.args.get('sheetName')
    if not sheet_name:
        return jsonify({"success": False, "message": "Sheet name is required"}), 400

    try:
        db = get_db()
        cursor = db.cursor(dictionary=True)

        # Check if table exists
        table_name = f"{sheet_name}_volume_report"
        cursor.execute("""
            SELECT COUNT(*) as table_exists 
            FROM information_schema.tables 
            WHERE table_schema = DATABASE() AND table_name = %s
        """, (table_name,))
        table_exists = cursor.fetchone()['table_exists']
        if not table_exists:
            return jsonify({"success": False, "message": "No volume report table found for this sheet"}), 404

        # Get all rows
        cursor.execute(f"""
            SELECT user_code, NN_count, OI_count, MS_count 
            FROM `{table_name}`
        """)
        rows = cursor.fetchall()

        if not rows:
            return jsonify({"success": False, "message": "No data found in volume report"}), 404

        # Totals
        total_nn = sum(row['NN_count'] or 0 for row in rows)
        total_oi = sum(row['OI_count'] or 0 for row in rows)
        total_ms = sum(row['MS_count'] or 0 for row in rows)

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Volume Report"

        # Headers
        headers = ["User Code","Input of Name & ID Number (NN)","Input ID Number (OI)","Change of Member Status (MS)"]

        ws.append(headers)

        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Rows
        for row in rows:
            ws.append([row['user_code'], row['NN_count'], row['OI_count'], row['MS_count']])

        # Totals row
        ws.append(["TOTAL", total_nn, total_oi, total_ms])
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=ws.max_row, column=col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')

        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            ws.column_dimensions[column_letter].width = (max_length + 2) * 1.2

        # Send file
        excel_io = BytesIO()
        wb.save(excel_io)
        excel_io.seek(0)

        return send_file(
            excel_io,
            as_attachment=True,
            download_name=f"Volume_Report_{sheet_name}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        app.logger.error(f"Excel report error: {str(e)}\n{traceback.format_exc()}")
        return jsonify({
            "success": False,
            "message": "Failed to generate Excel report",
            "error": str(e)
        }), 500
    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'db' in locals():
            db.close()

@app.route('/check-mfile-errors', methods=['GET'])
def check_mfile_errors():
    sheet_name = request.args.get('sheetName')
    if not sheet_name:
        return jsonify(success=False, message="Sheet name is required"), 400

    db = None
    cur = None
    try:
        db = get_db()
        cur = db.cursor(dictionary=True)
        
        table_name = f"{sheet_name}Mfile"
        
        # Check if table exists
        cur.execute("SHOW TABLES LIKE %s", (table_name,))
        if not cur.fetchone():
            return jsonify(success=False, message=f"Table {table_name} not found"), 404
        
        # Check for records with invalid lengths
        cur.execute(f"""
        SELECT COUNT(*) as error_count FROM `{table_name}`
        WHERE 
            LENGTH(batch_number) > 8 OR
            LENGTH(zone_code) > 1 OR
            LENGTH(emp_number) > 6 OR
            LENGTH(member_number) > 6 OR
            LENGTH(last_name) > 60 OR
            LENGTH(id_number) > 12 OR
            LENGTH(id_status) > 1 OR
            LENGTH(mem_status) > 1 OR
            LENGTH(operation_code) > 1 OR
            LENGTH(full_name) > 100
        """)
        
        result = cur.fetchone()
        error_count = result['error_count'] if result else 0
        
        return jsonify(
            success=True,
            errorCount=error_count,
            message=f"Found {error_count} records with formatting issues" if error_count > 0 
                   else "All records have valid formatting"
        )
        
    except Exception as e:
        return jsonify(success=False, message=str(e)), 500
    finally:
        if cur: cur.close()
        if db: db.close()
# === Get User Action Counts ===
# === Get User Action Counts ===
@app.route('/get-user-action-counts', methods=['GET'])
def get_user_action_counts():
    sheet_name = request.args.get('sheetName')
    try:
        db = get_db()
        cursor = db.cursor(dictionary=True)
        
        # Modified query to join with user table
        query = """
            SELECT 
                u.name as user_name,
                l.user_code,
                l.sheet_name,
                SUM(l.inserted) as total_inserted,
                SUM(l.updated) as total_updated,
                MAX(l.timestamp) as last_activity
            FROM user_action_log l
            LEFT JOIN user u ON l.user_code = u.user_code
        """
        
        params = []
        
        if sheet_name:
            query += " WHERE l.sheet_name = %s"
            params.append(sheet_name)
        
        query += " GROUP BY l.user_code, l.sheet_name, u.name"
        
        cursor.execute(query, params)
        counts = cursor.fetchall()
        
        return jsonify({"success": True, "data": counts})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500
    finally:
        cursor.close()
        db.close()

# === Generate Excel Report ===


# ... (your other Flask routes and setup code above) ...

@app.route('/generate-excel-report', methods=['POST'])
def generate_excel_report():
    db = None
    cursor = None
    try:
        data = request.get_json()
        if not data:
            return jsonify({"success": False, "message": "No data received"}), 400

        sheet_name = data.get('sheetName')
        if not sheet_name:
            return jsonify({"success": False, "message": "Sheet name is required"}), 400

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "User Activity Report"

        # Add headers
        headers = ["User Name", "User Code", "Sheet Name", "Inserted Rows", "Updated Rows", "Last Activity"]
        ws.append(headers)

        # Get data from database
        db = get_db()
        cursor = db.cursor(dictionary=True)
        
        query = """
            SELECT 
                COALESCE(u.name, l.user_code) as user_name,
                l.user_code,
                l.sheet_name,
                SUM(l.inserted) as total_inserted,
                SUM(l.updated) as total_updated,
                MAX(l.timestamp) as last_activity
            FROM user_action_log l
            LEFT JOIN user u ON l.user_code = u.user_code
            WHERE l.sheet_name = %s
            GROUP BY l.user_code, l.sheet_name, u.name
            ORDER BY last_activity DESC
        """
        
        cursor.execute(query, (sheet_name,))
        results = cursor.fetchall()

        if not results:
            return jsonify({"success": False, "message": "No data found for this sheet"}), 404

        # Add data rows
        for row in results:
            last_activity = row['last_activity'].strftime('%Y-%m-%d %H:%M:%S') if row['last_activity'] else 'N/A'
            ws.append([
                row['user_name'],
                row['user_code'],
                row['sheet_name'],
                row['total_inserted'],
                row['total_updated'],
                last_activity
            ])

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            ws.column_dimensions[column_letter].width = adjusted_width

        # Save to buffer
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(
            output,
            as_attachment=True,
            download_name=f"{sheet_name}_user_activity.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except mysql.connector.Error as db_err:
        return jsonify({
            "success": False,
            "message": f"Database error: {db_err.msg}",
            "error_code": db_err.errno
        }), 500
    except Exception as e:
        app.logger.error(f"Excel report error: {str(e)}\n{traceback.format_exc()}")
        return jsonify({
            "success": False,
            "message": "Failed to generate Excel report",
            "error": str(e)
        }), 500
    finally:
        if cursor:
            cursor.close()
        if db:
            db.close()

@app.route('/generate-pdf-report', methods=['POST'])
def generate_pdf_report():
    db = None
    cursor = None
    try:
        data = request.get_json()
        if not data:
            return jsonify({"success": False, "message": "No data received"}), 400

        sheet_name = data.get('sheetName')
        if not sheet_name:
            return jsonify({"success": False, "message": "Sheet name is required"}), 400

        # Create PDF
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)
        
        # Add title
        pdf.cell(200, 10, txt=f"User Activity Report - {sheet_name}", ln=1, align='C')
        pdf.ln(10)
        
        # Get data from database
        db = get_db()
        cursor = db.cursor(dictionary=True)
        
        query = """
            SELECT 
                COALESCE(u.name, l.user_code) as user_name,
                l.user_code,
                l.sheet_name,
                SUM(l.inserted) as total_inserted,
                SUM(l.updated) as total_updated,
                MAX(l.timestamp) as last_activity
            FROM user_action_log l
            LEFT JOIN user u ON l.user_code = u.user_code
            WHERE l.sheet_name = %s
            GROUP BY l.user_code, l.sheet_name, u.name
            ORDER BY last_activity DESC
        """
        
        cursor.execute(query, (sheet_name,))
        results = cursor.fetchall()

        if not results:
            return jsonify({"success": False, "message": "No data found for this sheet"}), 404

        # Set column widths
        col_widths = [40, 30, 30, 30, 30, 40]
        
        # Add table headers
        pdf.set_fill_color(200, 220, 255)
        headers = ["User Name", "User Code", "Sheet Name", "Inserted", "Updated", "Last Activity"]
        
        for i, header in enumerate(headers):
            pdf.cell(col_widths[i], 10, header, 1, 0, 'C', 1)
        pdf.ln()
        
        # Add data rows
        pdf.set_fill_color(255, 255, 255)
        for row in results:
            last_activity = row['last_activity'].strftime('%Y-%m-%d %H:%M') if row['last_activity'] else 'N/A'
            data = [
                row['user_name'],
                row['user_code'],
                row['sheet_name'],
                str(row['total_inserted']),
                str(row['total_updated']),
                last_activity
            ]
            
            for i, item in enumerate(data):
                pdf.cell(col_widths[i], 10, str(item), 1)
            pdf.ln()
        
        # Save to buffer
        buffer = BytesIO()
        pdf_bytes = pdf.output(dest='S').encode('latin-1')
        buffer.write(pdf_bytes)
        buffer.seek(0)
        
        return send_file(
            buffer,
            as_attachment=True,
            download_name=f"{sheet_name}_user_activity.pdf",
            mimetype="application/pdf"
        )
        
    except mysql.connector.Error as db_err:
        return jsonify({
            "success": False,
            "message": f"Database error: {db_err.msg}",
            "error_code": db_err.errno
        }), 500
    except Exception as e:
        app.logger.error(f"PDF report error: {str(e)}\n{traceback.format_exc()}")
        return jsonify({
            "success": False,
            "message": "Failed to generate PDF report",
            "error": str(e)
        }), 500
    finally:
        if cursor:
            cursor.close()
        if db:
            db.close()








        # Add these routes to your Flask app
@app.route('/admin-stats', methods=['GET'])
def get_admin_stats():
    try:
        db = get_db()
        cursor = db.cursor(dictionary=True)
        
        # Initialize stats with default values
        stats = {
            "totalUsers": 0,
            "activeToday": 0,
            "dataEntries": 0
        }

        # Get total users
        cursor.execute("SELECT COUNT(*) as total_users FROM user")
        total_users = cursor.fetchone()
        stats["totalUsers"] = total_users.get('total_users', 0) if total_users else 0

        # Get active users today
        today = datetime.datetime.now().strftime('%Y-%m-%d')
        cursor.execute("""
            SELECT COUNT(DISTINCT user_code) as active_today 
            FROM user_action_log 
            WHERE DATE(timestamp) = %s
        """, (today,))
        active_today = cursor.fetchone()
        stats["activeToday"] = active_today.get('active_today', 0) if active_today else 0

        # Get total data entries
        cursor.execute("SELECT SUM(inserted + updated) as total_entries FROM user_action_log")
        data_entries = cursor.fetchone()
        stats["dataEntries"] = data_entries.get('total_entries', 0) if data_entries else 0

        return jsonify({
            "success": True,
            "data": stats
        })

    except mysql.connector.Error as err:
        return jsonify({
            "success": False,
            "message": f"Database error: {err.msg}",
            "error_code": err.errno
        }), 500

    except Exception as e:
        return jsonify({
            "success": False,
            "message": f"Server error: {str(e)}"
        }), 500

    finally:
        if 'cursor' in locals():
            cursor.close()
        if 'db' in locals():
            db.close()

@app.route('/recent-activities', methods=['GET'])
def get_recent_activities():
    try:
        db = get_db()
        cursor = db.cursor(dictionary=True)
        
        cursor.execute("""
            SELECT 
                user_code,
                sheet_name,
                inserted,
                updated,
                timestamp,
                CONCAT(
                    CASE 
                        WHEN inserted > 0 THEN CONCAT(inserted, ' inserts to ', sheet_name)
                        ELSE CONCAT(updated, ' updates to ', sheet_name)
                    END,
                    ' by ', user_code
                ) as message,
                CASE 
                    WHEN inserted > 0 THEN 'success'
                    ELSE 'info'
                END as type,
                TIMESTAMPDIFF(MINUTE, timestamp, NOW()) as minutes_ago
            FROM user_action_log
            ORDER BY timestamp DESC
            LIMIT 10
        """)
        
        activities = []
        for row in cursor.fetchall():
            time_ago = ""
            if row['minutes_ago'] < 60:
                time_ago = f"{row['minutes_ago']} minutes ago"
            elif row['minutes_ago'] < 1440:
                hours = row['minutes_ago'] // 60
                time_ago = f"{hours} hour{'s' if hours > 1 else ''} ago"
            else:
                days = row['minutes_ago'] // 1440
                time_ago = f"{days} day{'s' if days > 1 else ''} ago"
                
            activities.append({
                "message": row['message'],
                "type": row['type'],
                "timestamp": row['timestamp'].isoformat(),
                "timeAgo": time_ago
            })
            
        return jsonify(activities)
        
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500
    finally:
        cursor.close()
        db.close()


if __name__ == "__main__":
    app.run(debug=True)


